import cv2
import pytesseract
import numpy as np
import os
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook
import re

# Optional if Tesseract is not in PATH
# pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

image_path = 'table.png'
excel_file = 'results.xlsx'
previous_source_file = 'previous_source.xlsx'

# Keep track of previously seen rows
previous_rows = set()
header_saved = False

# Create workbook if it doesn’t exist
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Result Data"
    wb.save(excel_file)
if not os.path.exists(previous_source_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR Source Data"
    wb.save(previous_source_file)
else:
    wb = load_workbook(previous_source_file)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        previous_rows.add(tuple(row))  # Full row tuple
    wb.close()

def clean_signal_time(signal_time_str):
    # Fix common OCR errors and insert space between date and time
    # Example fix: '2025-08-0723:45:00' -> '2025-08-07 23:45:00'
    fixed = re.sub(r"(\d{4}-\d{2}-\d{2})(\d{2}:\d{2}:\d{2})", r"\1 \2", signal_time_str)

    # Remove invalid seconds (e.g. 63, 65), clamp to 59 if needed
    # Extract hh:mm:ss part
    parts = fixed.split()
    if len(parts) == 2:
        time_part = parts[1]
        h, m, s = time_part.split(':')
        s = int(s)
        if s > 59:
            s = 59
        fixed = f"{parts[0]} {h}:{m}:{s:02d}"
    return fixed

def extract_table(image):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, 200, 255, cv2.THRESH_BINARY_INV)

    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
    h_lines = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, h_kernel)

    v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))
    v_lines = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, v_kernel)

    grid = cv2.add(h_lines, v_lines)
    contours, _ = cv2.findContours(grid, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    # Sort into rows
    def sort_contours(cnts):
        cnts = sorted(cnts, key=lambda c: cv2.boundingRect(c)[1])
        rows = []
        current_row = []
        last_y = None
        for c in cnts:
            x, y, w, h = cv2.boundingRect(c)
            if w < 20 or h < 15:
                continue
            if last_y is None or abs(y - last_y) < 10:
                current_row.append(c)
                last_y = y
            else:
                rows.append(sorted(current_row, key=lambda c: cv2.boundingRect(c)[0]))
                current_row = [c]
                last_y = y
        if current_row:
            rows.append(sorted(current_row, key=lambda c: cv2.boundingRect(c)[0]))
        return rows

    return sort_contours(contours)

def ocr_table(image, rows):
    table_data = []
    for row_i, row in enumerate(rows):
        row_texts = []
        for col_i, cell in enumerate(row):
            x, y, w, h = cv2.boundingRect(cell)
            roi = image[y:y+h, x:x+w]
            roi_gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
            _, roi_thresh = cv2.threshold(roi_gray, 200, 255, cv2.THRESH_BINARY_INV)
            text = pytesseract.image_to_string(
                roi_thresh,
                config='--psm 6 -c tessedit_char_whitelist=0123456789.:+-ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz'
            ).strip()
            # Clean SignalTime column (assume first column)
            if row_i > 0 and col_i == 0 and text:
                text = clean_signal_time(text)
            row_texts.append(text)
        table_data.append(row_texts)
    return table_data

def highlight_new_rows(image, rows, new_rows_data):
    for i, row in enumerate(rows):
        row_data = new_rows_data[i] if i < len(new_rows_data) else None
        for j, cell in enumerate(row):
            x, y, w, h = cv2.boundingRect(cell)
            if row_data and any(row_data):
                cv2.rectangle(image, (x, y), (x+w, y+h), (0, 255, 0), 2)  # Green for new
            else:
                cv2.rectangle(image, (x, y), (x+w, y+h), (128, 128, 128), 1)  # Gray for old
    return image

def append_to_excel(header, data_rows):
    global header_saved
    print(header, "header")
    print(data_rows, "data_rows")
    wb = load_workbook(excel_file)
    ws = wb.active

    if not header_saved:
        if ws.max_row == 1 and not any(ws.iter_rows(values_only=True)):
            init_header = header[0].split()
            init_header[0] = "SignalDate"
            update_header = init_header[:1] + ["SignalTime"] + init_header[1:9]
            ws.append(update_header)
        header_saved = True

    for row in data_rows:
        init_row = row[0].split()
        init_row[2] = init_row[2][:3] + "." + init_row[2][-3:]
        for i in range(4,len(init_row)):
            if "o" in init_row[i]:
                if(init_row[i][0] == "o"):
                    init_row[i] = init_row[i].replace("o", "9.")
                else:
                    init_row[i] = init_row[i].replace("o", "9")
        ws.append(init_row[0:10])

    wb.save(excel_file)
    wb.close()

def append_to_source_excel(header, data_rows):
    global header_saved
    print(header, "header")
    print(data_rows, "data_rows")
    wb = load_workbook(previous_source_file)
    ws = wb.active

    if not header_saved:
        if ws.max_row == 1 and not any(ws.iter_rows(values_only=True)):
            ws.append(header)
        header_saved = True

    for row in data_rows:
        ws.append(row)

    wb.save(previous_source_file)
    wb.close()
print("🔍 Watching for table updates... Press Ctrl+C to stop.")
try:
    while True:
        if os.path.exists(image_path):
            img = cv2.imread(image_path)
            table_rows = extract_table(img)
            table_data = ocr_table(img, table_rows)
            print(table_data, "111111")
            if not table_data:
                time.sleep(5)
                continue

            header = table_data[1]
            print(header, "222222")
            data_rows = table_data[2:]
            print(data_rows, "333333")
            
            # Filter out rows already in Excel
            new_rows = []
            print(new_rows, "444444")
            print(previous_rows, "555555")
            for r in data_rows:
                row_tuple = tuple(r)
                if row_tuple not in previous_rows:
                    previous_rows.add(row_tuple)
                    new_rows.append(r)
            print(new_rows, "666666")

            if new_rows:
                append_to_excel(header, new_rows)
                append_to_source_excel(header, new_rows)
                img = highlight_new_rows(img, table_rows[1:], new_rows)
                print(img, "777777")

            cv2.imshow("OCR Table Monitor", img)
            if cv2.waitKey(1) & 0xFF == 27:
                break

        time.sleep(5)

except KeyboardInterrupt:
    print("\n🛑 Stopped by user.")
finally:
    cv2.destroyAllWindows()
